"""
Excel Generation Service.

Generates Excel reports with pivoted student response data.
Uses pandas and openpyxl for Excel creation.
"""
import pandas as pd
from io import BytesIO
from typing import List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


class ExcelService:
    """
    Service for generating formatted Excel reports.
    """
    
    # Fixed student columns that appear first
    STUDENT_COLUMNS = [
        'TestDate',
        'SchoolId', 
        'SchoolName',
        'StudentId',
        'FirstName',
        'LastName',
        'Gender',
        'Grade',
        'Region'
    ]
    
    # Question detail columns (will be prefixed with Q1_, Q2_, etc.)
    QUESTION_COLUMNS = [
        'QuestionId',
        'Subject',
        'Level',
        'Type',
        'StudentAnswer',
        'CorrectAnswer',
        'Score'
    ]
    
    def __init__(self):
        # Styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def pivot_student_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Transform the raw query data (one row per student-question)
        into pivoted format (one row per student, question details as columns).
        
        Uses pandas pivot_table for fast O(n log n) performance.
        """
        if df.empty:
            return pd.DataFrame()
        
        # Get unique students with their base info
        student_cols = ['StudentId', 'FirstName', 'LastName', 'Gender', 'Grade', 
                       'SchoolId', 'SchoolName', 'Region', 'TestDate']
        
        # Get the student base data (drop duplicates)
        students_df = df[student_cols].drop_duplicates(subset=['StudentId'])
        
        # Format TestDate as readable string (YYYY-MM-DD)
        if 'TestDate' in students_df.columns:
            students_df = students_df.copy()
            students_df['TestDate'] = pd.to_datetime(students_df['TestDate']).dt.strftime('%Y-%m-%d')
        
        # Create a question number mapping (sorted by QuestionID)
        all_question_ids = sorted(df['QuestionID'].unique())
        q_id_to_num = {q_id: idx for idx, q_id in enumerate(all_question_ids, 1)}
        
        # Add question number to dataframe
        df = df.copy()
        df['QNum'] = df['QuestionID'].map(q_id_to_num)
        
        # Pivot each column we need using pivot_table (FAST!)
        pivot_cols = {
            'QuestionID': 'QuestionId',
            'Subject': 'Subject', 
            'Level': 'Level',
            'QuestionType': 'Type',
            'StudentAnswer': 'StudentAnswer',
            'CorrectAnswer': 'CorrectAnswer',
            'Score': 'Score'
        }
        
        # Create pivot for all columns at once
        pivoted = df.pivot_table(
            index='StudentId',
            columns='QNum',
            values=list(pivot_cols.keys()),
            aggfunc='first'
        )
        
        # Flatten column names: ('StudentAnswer', 1) -> 'Q1_StudentAnswer'
        new_columns = []
        for col in pivoted.columns:
            value_name, q_num = col
            new_col_name = f"Q{q_num}_{pivot_cols[value_name]}"
            new_columns.append(new_col_name)
        pivoted.columns = new_columns
        
        # Reset index to make StudentId a column
        pivoted = pivoted.reset_index()
        
        # Merge with student info
        result_df = students_df.merge(pivoted, on='StudentId', how='left')
        
        # Reorder columns: student info first, then question columns in order
        final_columns = self.STUDENT_COLUMNS.copy()
        for idx in range(1, len(all_question_ids) + 1):
            prefix = f"Q{idx}_"
            for col in self.QUESTION_COLUMNS:
                col_name = f"{prefix}{col}"
                if col_name in result_df.columns:
                    final_columns.append(col_name)
        
        # Only keep columns that exist
        final_columns = [c for c in final_columns if c in result_df.columns]
        result_df = result_df[final_columns]
        
        # Fill blanks for questions student didn't answer
        for col in result_df.columns:
            if col.endswith('_StudentAnswer'):
                result_df[col] = result_df[col].fillna('Not Answered')
            elif col.endswith('_Score'):
                result_df[col] = result_df[col].fillna(0)
            elif col.endswith('_QuestionId') or col.endswith('_Subject') or col.endswith('_Level') or col.endswith('_Type') or col.endswith('_CorrectAnswer'):
                result_df[col] = result_df[col].fillna('N/A')
        
        return result_df
    
    def generate_excel(
        self, 
        df: pd.DataFrame,
        contest_id: int,
        contest_info: Optional[dict] = None
    ) -> BytesIO:
        """
        Generate an Excel file from the pivoted DataFrame.
        Uses pandas to_excel for fast writing.
        Returns a BytesIO buffer containing the Excel file.
        """
        buffer = BytesIO()
        
        if df.empty:
            # Handle empty data case
            wb = Workbook()
            ws = wb.active
            ws['A1'] = f"Contest ID: {contest_id}"
            ws['A2'] = "No data found for the specified filters."
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        # Use pandas ExcelWriter for fast writing
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Write main data starting at row 4 (leave room for header)
            df.to_excel(writer, sheet_name='Student Responses', index=False, startrow=3)
            
            # Get the workbook and worksheet
            wb = writer.book
            ws = writer.sheets['Student Responses']
            
            # Add metadata header
            ws['A1'] = f"Contest ID: {contest_id}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Students: {len(df)} | Questions: {(len(df.columns) - len(self.STUDENT_COLUMNS)) // 7}"
            
            # Style the header row (row 4)
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=4, column=col_idx)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            
            # Freeze panes (header + student columns)
            ws.freeze_panes = ws.cell(row=5, column=len(self.STUDENT_COLUMNS) + 1)
            
            # Set reasonable column widths
            for i, col in enumerate(df.columns, 1):
                col_letter = get_column_letter(i)
                if col == 'TestDate':
                    ws.column_dimensions[col_letter].width = 12
                elif col == 'SchoolName':
                    ws.column_dimensions[col_letter].width = 25
                elif col in self.STUDENT_COLUMNS:
                    ws.column_dimensions[col_letter].width = 12
                else:
                    ws.column_dimensions[col_letter].width = 10
            
            # Add summary sheet
            summary_data = {
                'Metric': ['Contest ID', 'Total Students', 'Total Schools', 
                          'Total Questions', 'Generated At'],
                'Value': [
                    contest_id,
                    len(df),
                    df['SchoolId'].nunique() if 'SchoolId' in df.columns else 0,
                    (len(df.columns) - len(self.STUDENT_COLUMNS)) // 7,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        buffer.seek(0)
        return buffer


# Global service instance
excel_service = ExcelService()
